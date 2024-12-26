from datetime import datetime
from itertools import groupby

from App.model import TeacherPlan, Class, Subject, TeacherSubject, Semester,Profile, Student, Score, Exam, StudentClass
from App import db, app
from sqlalchemy.sql import func,case
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from datetime import datetime


def get_class_by_teacher_id(teacher_id):
    return (db.session.query(Class)
            .join(TeacherPlan, TeacherPlan.class_id.__eq__(Class.id))
            .join(TeacherSubject, TeacherSubject.id.__eq__(TeacherPlan.teacher_subject_id))
            .join(Semester, Semester.id == TeacherPlan.semester_id)
            .join(Subject, Subject.id == TeacherSubject.subject_id)
            .filter(TeacherSubject.teacher_id.__eq__(teacher_id)).all())


def get_plan_by_class_id(class_id,teacher_id):
    return (db.session.query(TeacherPlan)
            .join(Class, Class.id.__eq__(TeacherPlan.class_id))
            .join(TeacherSubject, TeacherSubject.id.__eq__(TeacherPlan.teacher_subject_id))
            .join(Semester, Semester.id == TeacherPlan.semester_id)
            .join(Subject, Subject.id == TeacherSubject.subject_id)
            .filter(TeacherPlan.class_id.__eq__(class_id),
                    TeacherSubject.teacher_id.__eq__(teacher_id),
                    Semester.year.__eq__(datetime.now().year)).all())


def get_plan(class_id,teacher_id,semester_id,subject_id):
    return (db.session.query(TeacherPlan)
            .join(Class, Class.id.__eq__(TeacherPlan.class_id))
            .join(TeacherSubject, TeacherSubject.id.__eq__(TeacherPlan.teacher_subject_id))
            .join(Semester, Semester.id == TeacherPlan.semester_id)
            .join(Subject, Subject.id == TeacherSubject.subject_id)
            .filter(TeacherPlan.class_id.__eq__(class_id),
                    TeacherPlan.semester_id.__eq__(semester_id),
                    TeacherSubject.teacher_id.__eq__(teacher_id),
                    TeacherSubject.subject_id.__eq__(subject_id)).first())


def get_teaching_plan_by_id(teacher_plan_id):
    return TeacherPlan.query.filter(TeacherPlan.id.__eq__(teacher_plan_id)).first()

def get_final_exam_scores(semester_id, subject_id, class_id):
    return (db.session.query(Student.id)
            .join(Exam, Exam.student_id == Student.id)
            .join(Score, Score.exam_id == Exam.id)
            .join(TeacherPlan, TeacherPlan.id == Exam.teacher_plan_id)
            .join(TeacherSubject, TeacherSubject.id == TeacherPlan.teacher_subject_id)
            .filter(
                TeacherPlan.semester_id == semester_id,
                TeacherSubject.subject_id == subject_id,
                TeacherPlan.class_id == class_id,
                Score.type_exam == "FINAL_EXAM"
            ).all())

def average_score(semester_id, subject_id, class_id):
    average = db.session.query(
        Student.id,
        Profile.name,
        func.sum(case(
            (Score.type_exam == "EXAM_15P", Score.value * 1),
            (Score.type_exam == "EXAM_45P", Score.value * 2),
            (Score.type_exam == "FINAL_EXAM", Score.value * 3)
        )) / func.sum(case(
            (Score.type_exam == "EXAM_15P", 1),
            (Score.type_exam == "EXAM_45P", 2),
            (Score.type_exam == "FINAL_EXAM", 3)
        )).label('average_score')
        ).select_from(Student
        ).join(Profile, Profile.id == Student.id
        ).join(StudentClass, StudentClass.student_id == Student.id
        ).join(Exam, Exam.student_id == Student.id
        ).join(Score, Score.exam_id == Exam.id
        ).join(TeacherPlan, TeacherPlan.id == Exam.teacher_plan_id
        ).join(TeacherSubject,TeacherSubject.id == TeacherPlan.teacher_subject_id
        ).filter(
            StudentClass.class_id == class_id,
            TeacherPlan.teacher_subjects.has(subject_id=subject_id),
            TeacherPlan.semester_id == semester_id,
        ).group_by(Student.id, Profile.name).all()
    return  average

def get_semesters_by_year(year):
    return (db.session.query(Semester)
            .filter(Semester.year.__eq__(year))
            .all())

def init_student_scores(teach_plan):
    student_scores = {}
    for student_class in teach_plan.classes.student_class:
        student_scores[student_class.students.id] = {'HK1': None, 'HK2': None}
    return student_scores


def update_scores(semester, averages, student_final_check, student_scores):
    for student_id, name, avg_score in averages:
        if student_id in student_final_check and avg_score is not None:
            semester_name = f"{semester.name.name}"
            student_scores[student_id][semester_name] = round(float(avg_score), 2)


def process_scores(semester, teach_plan, student_scores):
    students_final = get_final_exam_scores(
        semester.id,
        teach_plan.teacher_subjects.subject_id,
        teach_plan.class_id
    )
    student_final_check = {student.id for student in students_final}

    if student_final_check:
        averages = average_score(
            semester.id,
            teach_plan.teacher_subjects.subject_id,
            teach_plan.class_id
        )
        update_scores(semester, averages, student_final_check, student_scores)


def export_score(teach_plan, get_score):
    wb = Workbook()
    ws = wb.active
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ws.merge_cells('A1:E1')
    ws['A1'] = 'BẢNG ĐIỂM MÔN HỌC'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = f'Lớp: {teach_plan.classes.name}'
    ws['D2'] = f'Môn: {teach_plan.teacher_subjects.subjects.name}'
    ws.merge_cells('A2:B2')
    ws.merge_cells('D2:E2')

    ws['A3'] = f'Học kỳ: {teach_plan.semester.name.name}'
    ws['D3'] = f'Năm học: {teach_plan.semester.year}'
    ws.merge_cells('A3:B3')
    ws.merge_cells('D3:E3')

    headers = ['STT', 'Họ Tên', 'Điểm 15p', 'Điểm 1 Tiết', 'Điểm Thi']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    current_row = 5 #Bắt đầu từ dòng 5
    for idx, student_class in enumerate(teach_plan.classes.student_class, 1):
        st = student_class.students

        ws.cell(row=current_row, column=1, value=idx)
        ws.cell(row=current_row, column=2, value=st.profile.name)

        diem_15p = []
        for i in range(teach_plan.teacher_subjects.subjects.number_of_15p):
            score = get_score(teach_plan.id, st.id, 'EXAM_15P', i + 1)
            if score and score.value is not None:
                diem_15p.append(str(score.value))
        ws.cell(row=current_row, column=3, value=' | '.join(diem_15p))

        diem_45p = []
        for i in range(teach_plan.teacher_subjects.subjects.number_of_45p):
            score = get_score(teach_plan.id, st.id, 'EXAM_45P', i + 1)
            if score and score.value is not None:
                diem_45p.append(str(score.value))
        ws.cell(row=current_row, column=4, value=' | '.join(diem_45p))

        score_final = get_score(teach_plan.id, st.id, 'FINAL_EXAM', 1)
        if score_final and score_final.value is not None:
            ws.cell(row=current_row, column=5, value=score_final.value)

        #Thêm viền và căn giữa cho tất cả các ô
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        current_row += 1

    column_widths = {
        'A': 8,
        'B': 30,
        'C': 25,
        'D': 25,
        'E': 15
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    return wb



if __name__ == '__main__':
    with app.app_context():
        print(average_score(subject_id=1,semester_id=5,class_id=1))
